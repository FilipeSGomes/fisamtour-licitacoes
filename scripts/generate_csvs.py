#!/usr/bin/env python3
"""Gera CSVs iniciais em appScript/csv/ a partir de LICITACÕES.xlsx."""

import csv
import re
import uuid
from datetime import UTC, datetime, timedelta, time
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
CSV_DIR = ROOT / "appScript" / "csv"
XLSX = ROOT / "LICITACÕES.xlsx"

LICITACOES = [
    ("4-rcc-rs", "4º RCC - RS", "passagens"),
    ("6-rcc-rs", "6º RCC - RS", "passagens"),
    ("13-rcm-sp", "13º RCM - SP", "passagens"),
    ("cais-lins-sp", "CAIS LINS - SP", "passagens"),
    ("tj-vitoria-es", "TJ VITORIA - ES", "hospedagem"),
    ("coren-ma", "COREN - MA", "veiculo"),
    ("tunas-pr", "TUNAS - PR", "calculadora"),
]

SHEET_TO_ID = {
    "4º RCC - RS": "4-rcc-rs",
    "6º RCC - RS": "6-rcc-rs",
    "13º RCM - SP": "13-rcm-sp",
    "CAIS LINS - SP": "cais-lins-sp",
    "TJ VITORIA - ES": "tj-vitoria-es",
    "COREN -MA": "coren-ma",
    "TUNAS - PR": "tunas-pr",
}


def new_id(prefix=""):
    return prefix + str(uuid.uuid4().int)[:12]


def excel_date(v):
    if v is None or v == "":
        return ""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, time):
        return ""
    s = str(v).strip()
    if re.match(r"^\d{1,2}/\d", s):
        for fmt in ("%d/%m/%Y", "%d/%m%Y"):
            try:
                return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
            except ValueError:
                pass
    try:
        n = float(v)
        if 30000 < n < 60000:
            return (datetime(1899, 12, 30) + timedelta(days=n)).strftime("%Y-%m-%d")
    except (TypeError, ValueError):
        pass
    return s


def excel_time(v):
    if v is None or v == "":
        return ""
    if isinstance(v, time):
        return v.strftime("%H:%M")
    if isinstance(v, datetime):
        return v.strftime("%H:%M")
    try:
        n = float(v)
        if 0 <= n < 1:
            total_min = int(round(n * 24 * 60))
            h, m = divmod(total_min, 60)
            return f"{h:02d}:{m:02d}"
    except (TypeError, ValueError):
        pass
    return str(v).strip()


def num(v):
    if v is None or v == "":
        return ""
    try:
        return round(float(v), 2)
    except (TypeError, ValueError):
        return ""


def competencia_from_date(date_str):
    return date_str[:7] if date_str and len(date_str) >= 7 else ""


def is_total_row(values):
    joined = " ".join(str(x or "").upper() for x in values)
    if "TOTAL" in joined:
        return True
    first = str(values[0] or "").strip().upper()
    months = (
        "JANEIRO", "FEVEREIRO", "MARÇO", "MARCO", "ABRIL", "MAIO", "JUNHO",
        "JULHO", "AGOSTO", "SETEMBRO", "OUTUBRO", "NOVEMBRO", "DEZEMBRO",
    )
    return first in months


def write_csv(name, headers, rows):
    CSV_DIR.mkdir(parents=True, exist_ok=True)
    path = CSV_DIR / name
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(headers)
        w.writerows(rows)
    print(f"{name}: {len(rows)} linhas")


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    lic_nome_map = {i: n for i, n, _ in LICITACOES}

    write_csv(
        "Catalogo_Licitacoes.csv",
        ["id", "nome", "tipo", "status"],
        [[i, n, t, "ativo"] for i, n, t in LICITACOES],
    )
    write_csv("Catalogo_Fornecedores.csv", ["id", "nome", "status"], [["op-1", "Operadora", "ativo"]])
    write_csv(
        "Catalogo_Tarifas.csv",
        ["id", "licitacao_id", "codigo", "nome", "valor", "custo", "editavel", "status"],
        [
            ["t-coren-carro", "coren-ma", "carro", "Carro", 481.76, 143, "false", "ativo"],
            ["t-coren-ped", "coren-ma", "pedestre", "Pedestre", 32.12, 11, "false", "ativo"],
            ["t-tj-diaria", "tj-vitoria-es", "diaria", "Hospedagem (diária)", 369.87, "", "false", "ativo"],
            ["t-tj-ref", "tj-vitoria-es", "refeicao", "Alimentação", 189.98, "", "false", "ativo"],
            ["t-tunas-livre", "tunas-pr", "livre", "Valor livre", "", "", "true", "ativo"],
        ],
    )
    write_csv(
        "Ordens_Servico.csv",
        [
            "id", "competencia", "data", "licitacao_id", "licitacao_nome", "tarifa_codigo", "tarifa_nome",
            "fornecedor", "descricao", "valor", "custo", "lucro", "status_os", "pago", "comprovante_url",
            "faturado_em", "created_at", "updated_at",
        ],
        [],
    )
    write_csv("Fechamentos.csv", ["competencia", "total_receitas", "total_despesas", "saldo", "fechado_em"], [])

    pass_rows = []
    ws = wb["Página8"]
    for r in range(2, ws.max_row + 1):
        nome = ws.cell(r, 1).value
        if not nome or not str(nome).strip():
            continue
        pass_rows.append([
            new_id("p-"),
            str(nome).strip(),
            excel_date(ws.cell(r, 2).value),
            str(ws.cell(r, 3).value or "").strip(),
            "ativo",
        ])
    write_csv("passageiros.csv", ["id", "nome", "data_nasc", "cpf", "status"], pass_rows)

    passagem_rows = []
    for sheet in ["6º RCC - RS", "13º RCM - SP"]:
        ws = wb[sheet]
        lic_id = SHEET_TO_ID[sheet]
        for r in range(2, ws.max_row + 1):
            row = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
            if not row[0] or is_total_row(row):
                continue
            colab = row[10]
            if not colab or "TOTAL" in str(colab).upper():
                continue
            valor = num(row[13]) if len(row) > 13 else ""
            custo = num(row[14]) if len(row) > 14 else ""
            lucro = num(row[15]) if len(row) > 15 else ""
            if valor != "" and custo != "" and lucro == "":
                lucro = round(float(valor) - float(custo), 2)
            passagem_rows.append([
                new_id("pg-"), lic_id,
                excel_date(row[0]), excel_date(row[1]), excel_date(row[2]),
                str(row[3] or "").strip(), str(row[4] or "").strip(),
                excel_time(row[5]), excel_time(row[6]),
                num(row[7]), str(row[8] or "").strip(), str(row[9] or "").strip(),
                str(colab).strip(), str(row[11] or "").strip(), excel_date(row[12]),
                valor, custo, lucro, num(row[16]) if len(row) > 16 else "",
                str(row[17] or "").strip() if len(row) > 17 else "", "ativo",
            ])

    ws = wb["CAIS LINS - SP"]
    lic_id = SHEET_TO_ID["CAIS LINS - SP"]
    for r in range(2, ws.max_row + 1):
        row = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if not row[0] or is_total_row(row):
            continue
        servidor = row[5]
        if not servidor or "TOTAL" in str(servidor).upper():
            continue
        data_ida = excel_date(row[2])
        valor = num(row[9])
        custo = num(row[10])
        lucro = num(row[11])
        if valor != "" and custo != "" and lucro == "":
            lucro = round(float(valor) - float(custo), 2)
        passagem_rows.append([
            new_id("pg-"), lic_id,
            excel_date(row[0]), data_ida, data_ida,
            str(row[3] or "").strip(), str(row[4] or "").strip(),
            "", "", "", str(row[1] or "").strip(), "",
            str(servidor).strip(), str(row[7] or "").strip(), excel_date(row[8]),
            valor, custo, lucro, "", str(row[13] or "").strip(), "ativo",
        ])

    write_csv(
        "reg_passagens.csv",
        [
            "id", "licitacao_id", "data_solicitacao", "data_ida", "data_chegada", "de", "para",
            "horario_saida", "horario_chegada", "assento", "bilhete", "classe", "colaborador", "cpf",
            "dt_nasc", "valor", "custo", "lucro", "percentil", "pago", "status",
        ],
        passagem_rows,
    )

    hosp_rows = []
    ws = wb["TJ VITORIA - ES"]
    lic_id = SHEET_TO_ID["TJ VITORIA - ES"]
    for r in range(2, ws.max_row + 1):
        row = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if not row[0] or is_total_row(row):
            continue
        colab = row[4]
        if not colab or "TOTAL" in str(colab).upper():
            continue
        vlr_total = num(row[8])
        if vlr_total == "" and str(row[7] or "").upper() == "TOTAL":
            continue
        hosp_rows.append([
            new_id("hp-"), lic_id,
            excel_date(row[0]), excel_date(row[1]), excel_date(row[2]), str(row[3] or "").strip(),
            str(colab).strip(), num(row[5]), num(row[6]), num(row[7]), vlr_total,
            num(row[9]), num(row[10]), num(row[11]), str(row[12] or "").strip(),
            str(row[13] or "").strip(), str(row[14] or "").strip(), "nao", "ativo",
        ])
    write_csv(
        "reg_hospedagem.csv",
        [
            "id", "licitacao_id", "dt_solicitacao", "check_in", "check_out", "diarias", "colaborador",
            "refeicoes", "vlr_diaria", "prev_vlr_refeicao", "vlr_total", "custo_hospedagem",
            "custo_refeicao", "custo_tt", "lucro", "prev_pgto", "enviado_faturamento", "pago", "status",
        ],
        hosp_rows,
    )

    veic_rows = []
    ws = wb["COREN -MA"]
    lic_id = SHEET_TO_ID["COREN -MA"]
    for r in range(2, ws.max_row + 1):
        row = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if not row[0] or is_total_row(row):
            continue
        colab = row[5]
        if not colab or "TOTAL" in str(colab).upper():
            continue
        valor = num(row[8])
        custo = num(row[9])
        lucro = num(row[10])
        if valor != "" and custo != "" and lucro == "":
            lucro = round(float(valor) - float(custo), 2)
        veic_rows.append([
            new_id("vc-"), lic_id,
            excel_date(row[0]), str(row[1] or "").strip(), str(row[2] or "").strip(),
            str(row[3] or "").strip(), excel_date(row[4]),
            str(colab).strip(), str(row[6] or "").strip(), excel_date(row[7]),
            valor, custo, lucro, str(row[11] or "").strip(), str(row[12] or "").strip(), "nao", "ativo",
        ])
    write_csv(
        "reg_veiculo.csv",
        [
            "id", "licitacao_id", "data_solicitacao", "modelo_veiculo", "placa", "saida", "data",
            "colaborador", "cpf", "dt_nascimento", "valor", "custo", "lucro", "prev_pgto",
            "enviado_faturamento", "pago", "status",
        ],
        veic_rows,
    )

    now = datetime.now(UTC).strftime("%Y-%m-%dT%H:%M:%S.000Z")
    lanc_rows = []

    def add_lanc(origem, origem_id, lic_id, data, tipo, categoria, descricao, valor, custo=""):
        if valor == "" or valor is None:
            return
        try:
            v = float(valor)
            if v == 0:
                return
        except (TypeError, ValueError):
            return
        comp = competencia_from_date(data)
        if not comp:
            return
        c = ""
        if custo != "" and custo is not None:
            try:
                c = float(custo)
            except (TypeError, ValueError):
                c = ""
        lanc_rows.append([
            new_id("lc-"), comp, data, lic_nome_map.get(lic_id, lic_id), "em_andamento",
            tipo, categoria, "", descricao, v, c if tipo == "despesa" else (c if c != "" else ""),
            "", origem, origem_id, now, now,
        ])

    for row in passagem_rows:
        rid, lic_id = row[0], row[1]
        data = row[3] or row[2]
        desc = f"{row[5]} → {row[6]} | {row[12]}".strip(" |")
        add_lanc("reg_passagens", rid, lic_id, data, "receita", "passagens", desc, row[15], "")
        add_lanc("reg_passagens", rid, lic_id, data, "despesa", "passagens", f"Custo: {desc}", row[16], row[16])

    for row in hosp_rows:
        rid, lic_id = row[0], row[1]
        data = row[3] or row[2]
        desc = f"Hospedagem | {row[6]}"
        add_lanc("reg_hospedagem", rid, lic_id, data, "receita", "hospedagem", desc, row[10], "")
        custo = row[13] or row[11]
        add_lanc("reg_hospedagem", rid, lic_id, data, "despesa", "hospedagem", f"Custo: {desc}", custo, custo)

    for row in veic_rows:
        rid, lic_id = row[0], row[1]
        data = row[6] or row[2]
        desc = f"{row[5]} | {row[7]}"
        add_lanc("reg_veiculo", rid, lic_id, data, "receita", "veiculo", desc, row[10], "")
        add_lanc("reg_veiculo", rid, lic_id, data, "despesa", "veiculo", f"Custo: {desc}", row[11], row[11])

    write_csv(
        "Lancamentos.csv",
        [
            "id", "competencia", "data", "licitacao", "status", "tipo", "categoria", "fornecedor",
            "descricao", "valor", "custo", "comprovante_url", "origem", "origem_id", "created_at", "updated_at",
        ],
        lanc_rows,
    )

    from collections import Counter
    print("Lancamentos por competência:", dict(sorted(Counter(r[1] for r in lanc_rows).items())))


if __name__ == "__main__":
    main()
